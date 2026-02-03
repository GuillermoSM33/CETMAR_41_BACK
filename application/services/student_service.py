import io
import re
from typing import Any, Dict, List, Optional

import bcrypt
from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session
from sqlalchemy.orm import joinedload

from application.dtos.students.get_student_dto import GetStudentDetailDTO
from application.dtos.students.update_student_dto import UpdateStudentDTO
from infrastructure.persistence.models.auth_model import AuthModel
from infrastructure.persistence.models.identity_model import IdentityModel
from infrastructure.persistence.models.role_model import RoleModel
from infrastructure.persistence.models.user_model import UserModel

def _hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def _to_str(v: Any) -> str:
    """Normalize excel cell value to safe string."""
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v).strip()


def _to_phone_int(v: Any) -> int:
    s = _to_str(v)
    if not s:
        return 0
    digits = re.sub(r"\D", "", s)
    if not digits:
        return 0
    try:
        return int(digits)
    except Exception:
        return 0

def import_students_from_excel(
    db: Session,
    file_bytes: bytes,
    create_auth: bool = False,
    default_password: str = "123456",
    sheet_name: Optional[str] = None,
) -> List[Dict[str, Any]]:

    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [_to_str(cell.value) for cell in ws[1]]

    results: List[Dict[str, Any]] = []

    student_role = db.query(RoleModel).filter(RoleModel.Role_Name == "Student").first()
    if not student_role:
        student_role = RoleModel(Role_Name="Student")
        db.add(student_role)
        db.commit()

    for row_values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, row_values))

        row_norm = {
            (k or "").strip().lower(): _to_str(v)
            for k, v in row.items()
        }

        nombres = row_norm.get("nombres") or row_norm.get("names")
        apellido_paterno = row_norm.get("apellido paterno") or row_norm.get("apellido_paterno")
        apellido_materno = row_norm.get("apellido materno") or row_norm.get("apellido_materno")
        correo = row_norm.get("correo") or row_norm.get("email")
        correo_l = correo.lower().strip() if correo else None

        numero_control = (
            row_norm.get("número de control")
            or row_norm.get("numero de control")
            or row_norm.get("numero_control")
        )

        curp = row_norm.get("curp")
        grupo = row_norm.get("grupo")
        carrera = row_norm.get("carrera")
        horario = row_norm.get("horario")

        telephone_default = row_norm.get("teléfono") or row_norm.get("telefono")
        telephone_value = _to_phone_int(telephone_default)

        created_identity_id: Optional[int] = None
        created_user_id: Optional[int] = None
        status = "skipped"

        identity: Optional[IdentityModel] = None
        if numero_control:
            identity = (
                db.query(IdentityModel)
                .filter(IdentityModel.Student_Control_Number == numero_control)
                .first()
            )
        if not identity and curp:
            identity = db.query(IdentityModel).filter(IdentityModel.CURP == curp).first()

        try:
            if not identity:
                identity = IdentityModel(
                    Student_Control_Number=numero_control or f"unknown-{correo_l or ''}",
                    CURP=curp or None,
                    Full_Name=f"{nombres or ''} {apellido_paterno or ''} {apellido_materno or ''}".strip() or None,
                    Grupo=grupo or None,
                    Schedule=horario or None,
                    Major=carrera or None,
                )
                db.add(identity)
                db.flush()
                created_identity_id = identity.Id
            else:
                updated_identity = False

                full_name = f"{nombres or ''} {apellido_paterno or ''} {apellido_materno or ''}".strip() or None

                if numero_control and identity.Student_Control_Number != numero_control:
                    identity.Student_Control_Number = numero_control
                    updated_identity = True

                if curp and identity.CURP != curp:
                    identity.CURP = curp
                    updated_identity = True

                if grupo and identity.Grupo != grupo:
                    identity.Grupo = grupo
                    updated_identity = True

                if horario and identity.Schedule != horario:
                    identity.Schedule = horario
                    updated_identity = True

                if carrera and identity.Major != carrera:
                    identity.Major = carrera
                    updated_identity = True

                if full_name and (not identity.Full_Name or identity.Full_Name.strip() != full_name):
                    identity.Full_Name = full_name
                    updated_identity = True

                if updated_identity:
                    db.add(identity)
                    db.flush()

            if correo_l:
                user = (
                    db.query(UserModel)
                    .filter(func.lower(UserModel.User_Email) == correo_l)
                    .first()
                )

                full_name_user = f"{nombres or ''} {apellido_paterno or ''} {apellido_materno or ''}".strip() or correo_l

                if not user:
                    user = UserModel(
                        User_Name=full_name_user,
                        User_Email=correo_l,
                        FK_Rol_ID=student_role.Id,
                        Telephone=telephone_value,
                        FK_Identity_ID=identity.Id,
                    )
                    db.add(user)
                    db.flush()
                    created_user_id = user.Id

                    if create_auth:
                        hashed = _hash_password(default_password)
                        auth = AuthModel(FK_User_ID=user.Id, Hashed_Password=hashed)
                        db.add(auth)
                        db.flush()
                else:
                    updated_user = False

                    if telephone_value and (not user.Telephone or int(user.Telephone) != int(telephone_value)):
                        user.Telephone = telephone_value
                        updated_user = True

                    if user.FK_Identity_ID != identity.Id:
                        user.FK_Identity_ID = identity.Id
                        updated_user = True

                    if user.FK_Rol_ID != student_role.Id:
                        user.FK_Rol_ID = student_role.Id
                        updated_user = True

                    if full_name_user and (not user.User_Name or user.User_Name.strip() != full_name_user):
                        user.User_Name = full_name_user
                        updated_user = True

                    if updated_user:
                        db.add(user)
                        db.flush()

                    created_user_id = user.Id

            db.commit()
            status = "created" if (created_identity_id or created_user_id) else "exists"

        except Exception as e:
            db.rollback()
            results.append({"row": row, "status": "error", "error": str(e)})
            continue

        results.append(
            {
                "row": row,
                "status": status,
                "identity_id": created_identity_id or (identity.Id if identity else None),
                "user_id": created_user_id,
            }
        )

    return results

def get_all_students_service(db: Session, role_name: str) -> List[GetStudentDetailDTO]:
    
    # 1. Buscar el Role ID
    role = db.query(RoleModel).filter(func.lower(RoleModel.Role_Name) == role_name.lower()).first()
    if not role:
        return []
    
    # 2. Cargar Usuarios y su Identity en una sola consulta (JOINEDLOAD)
    # Esto es vital para evitar problemas de rendimiento (consultas N+1)
    users_with_identity = db.query(UserModel).options(
        joinedload(UserModel.identity) # Carga ansiosa de la relación 'identity'
    ).filter(
        UserModel.FK_Rol_ID == role.Id
    ).all()
    
    # 3. Mapear Modelos a DTOs
    result_dtos: List[GetStudentDetailDTO] = []
    
    for user in users_with_identity:
        # La relación 'identity' se carga gracias al joinedload
        identity_data = user.identity
        
        # Mapeo Manual: Pasamos los campos de UserModel y los de IdentityModel
        dto = GetStudentDetailDTO(
            # Campos de UserModel (Base DTO, incluyendo las FKs obligatorias)
            Id=user.Id,
            User_Name=user.User_Name,
            User_Email=user.User_Email,
            Telephone=user.Telephone,
            FK_Rol_ID=user.FK_Rol_ID,
            FK_Identity_ID=user.FK_Identity_ID,
            
            # Campos de IdentityModel (Añadidos en GetStudentDetailDTO)
            # Nota: Si los nombres de los campos en IdentityModel son diferentes, ajusta aquí:
            Matricula=None,
            Numero_Control=identity_data.Student_Control_Number if identity_data else None,
            CURP=identity_data.CURP if identity_data else None,
            
            # Grupo / Carrera
            Grupo=identity_data.Grupo if identity_data and hasattr(identity_data, 'Grupo') else None,
            Carrera=identity_data.Major if identity_data and hasattr(identity_data, 'Major') else None,
        )
        result_dtos.append(dto)

    return result_dtos

def update_student_service(db: Session, user_id: int, student_data: UpdateStudentDTO) -> UserModel:
    student = db.query(UserModel).filter(UserModel.Id == user_id).first()
    if not student:
        raise KeyError("Usuario no encontrado")

    student.User_Name = student_data.User_Name
    student.User_Email = student_data.User_Email
    student.FK_Rol_ID = student_data.FK_Rol_ID
    student.Telephone = int(student_data.Telephone)
    student.FK_Identity_ID = student_data.FK_Identity_ID

    db.commit()
    db.refresh(student)
    return student